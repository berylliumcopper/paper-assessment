"""Build compact analysis context from extracted paper artifacts."""

from __future__ import annotations

import json
import subprocess
import sys
import tempfile
from dataclasses import dataclass, field
from pathlib import Path

from assessment.source_data_availability import (
    AUDIT_FILENAME,
    extract_archive_if_needed,
    find_info_files_in_member_list,
    format_audit_for_prompt,
    read_audit_file,
    scan_availability_statement,
    load_combined_article_markdown,
)

# Snippet sizing for supplementary text files (whole file vs preview for the model).
SUPP_TEXT_FULL_MAX_CHARS = 24_000
SUPP_TEXT_PREVIEW_MAX_CHARS = 8_000
SUPP_TEXT_READ_CAP_BYTES = 2 * 1024 * 1024

# Total character budgets to stay within model context limits (~262K tokens).
# Converted PDF markdown (main article from Marker) — roughly 150K chars ≈ 38K tokens.
MAX_CONVERTED_CHARS = 150_000
# Total of all supplementary text snippets combined — 120K chars ≈ 34K tokens.
MAX_SUPP_SNIPPET_TOTAL_CHARS = 120_000
# Data-file analysis selected-file previews — 60K chars ≈ 15K tokens.
MAX_DATA_ANALYSIS_SNIPPET_TOTAL_CHARS = 60_000

# Supplementary PDFs are converted with Marker to _readable/*.md; that markdown is included in the supplementary snippet budget.


def _truncate(text: str, max_chars: int | None) -> str:
    if max_chars is None:
        return text
    if len(text) <= max_chars:
        return text
    return f"{text[:max_chars]}\n\n[TRUNCATED {len(text) - max_chars} CHARS]"


def _text_file_snippet_for_model(path: Path, *, max_full: int, max_preview: int, read_cap_bytes: int) -> str:
    """Prefer full text for small files; larger files get a capped read and preview-only body."""
    try:
        nbytes = path.stat().st_size
    except OSError:
        return ""
    try:
        to_read = min(nbytes, read_cap_bytes) if nbytes else read_cap_bytes
        with path.open("rb") as f:
            raw = f.read(to_read)
        text = raw.decode("utf-8", errors="ignore")
    except OSError:
        return ""
    if nbytes <= max_full and len(text) <= max_full:
        return text
    if len(text) <= max_full:
        return text + f"\n\n[NOTE: on-disk size {nbytes} bytes exceeds inline full limit {max_full}; preview may be partial.]"
    preview = text[:max_preview]
    tail = f"\n\n[PREVIEW ONLY on_disk_bytes={nbytes} preview_chars={max_preview}]"
    return preview + tail


def _extract_pdf_text_snippet(path: Path, max_chars: int | None) -> str:
    """Best-effort PDF extraction using Marker first, then pypdf fallback. ``max_chars=None`` = no truncation."""
    marker_text = _extract_pdf_text_with_marker(path, max_chars)
    if marker_text:
        return marker_text

    try:
        from pypdf import PdfReader  # type: ignore
    except Exception:
        return "PDF text extraction unavailable (Marker/pypdf not available)."

    try:
        reader = PdfReader(str(path))
        chunks: list[str] = []
        # Limit pages to keep prompt size manageable.
        for page in reader.pages[:8]:
            text = page.extract_text() or ""
            if text.strip():
                chunks.append(text.strip())
        combined = "\n\n".join(chunks).strip()
        if not combined:
            return "PDF text extraction returned no readable text."
        return _truncate(combined, max_chars)
    except Exception as exc:  # noqa: BLE001
        return f"PDF text extraction failed: {exc}"


def _extract_pdf_text_with_marker(path: Path, max_chars: int | None) -> str | None:
    """Try extracting PDF text via marker_single if available. ``max_chars=None`` = full markdown."""
    try:
        with tempfile.TemporaryDirectory(prefix="supp_pdf_marker_") as tmpdir:
            out_dir = Path(tmpdir)
            command = [
                "marker_single",
                str(path),
                "--output_dir",
                str(out_dir),
            ]
            completed = subprocess.run(command, capture_output=True, text=True, check=False)
            if completed.returncode != 0:
                return None

            md_candidates = sorted(out_dir.rglob("*.md"))
            if not md_candidates:
                return None
            text = md_candidates[0].read_text(encoding="utf-8", errors="ignore").strip()
            if not text:
                return "Marker extraction produced empty markdown."
            return _truncate(text, max_chars)
    except Exception:
        return None


def _extract_excel_text_snippet(path: Path, max_chars: int) -> tuple[str, list[str]]:
    """Best-effort extraction from xls/xlsx with sheet metadata notes."""
    notes: list[str] = []
    try:
        import pandas as pd  # type: ignore

        sheets = pd.read_excel(path, sheet_name=None)
        sections: list[str] = []
        for sheet_name, df in sheets.items():
            row_count = int(df.shape[0])
            col_count = int(df.shape[1])
            total_cells = max(1, row_count * max(1, col_count))
            missing_cells = int(df.isna().sum().sum())
            missing_ratio = missing_cells / total_cells
            notes.append(
                f"{path.name}::{sheet_name} rows={row_count} cols={col_count} missing={missing_cells} ({missing_ratio:.1%})"
            )
            preview = df.head(30).fillna("").to_csv(index=False)
            sections.append(f"## Sheet: {sheet_name}\n{preview}")
        merged = "\n\n".join(sections).strip()
        if not merged:
            return "Spreadsheet appears empty.", notes
        return _truncate(merged, max_chars), notes
    except Exception:
        pass

    # Fallback for .xlsx if openpyxl is available.
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=path, read_only=True, data_only=True)
            sections: list[str] = []
            for ws in wb.worksheets[:5]:
                rows: list[str] = []
                row_count = 0
                max_cols = 0
                non_empty_cells = 0
                for row in ws.iter_rows(min_row=1, max_row=30, values_only=True):
                    vals = ["" if v is None else str(v) for v in row]
                    row_count += 1
                    max_cols = max(max_cols, len(vals))
                    non_empty_cells += sum(1 for v in vals if v != "")
                    if any(vals):
                        rows.append("\t".join(vals))
                sampled_total = max(1, row_count * max(1, max_cols))
                missing_cells = max(0, sampled_total - non_empty_cells)
                notes.append(
                    f"{path.name}::{ws.title} sampled_rows={row_count} sampled_cols={max_cols} sampled_missing={missing_cells} ({missing_cells / sampled_total:.1%})"
                )
                if rows:
                    sections.append(f"## Sheet: {ws.title}\n" + "\n".join(rows))
            merged = "\n\n".join(sections).strip()
            if not merged:
                return "Spreadsheet appears empty.", notes
            return _truncate(merged, max_chars), notes
        except Exception as exc:  # noqa: BLE001
            return f"Spreadsheet extraction failed: {exc}", notes

    return "Spreadsheet extraction unavailable (install pandas/openpyxl).", notes


def _safe_name(value: str) -> str:
    cleaned = "".join(ch if ch.isalnum() or ch in {"-", "_", "."} else "_" for ch in value).strip("_.")
    return cleaned or "sheet"


def _archive_stem_for_extract_dir(archive_path: Path) -> str:
    name = archive_path.name
    lower = name.lower()
    if lower.endswith(".tar.gz"):
        return _safe_name(name[: -len(".tar.gz")])
    return _safe_name(archive_path.stem)


def _ensure_archive_extracted(archive_path: Path, supplementary_dir: Path) -> tuple[Path, list[str], list[str]]:
    """
    Extract zip/tar.gz into supplementary/_extracted/<stem>/.
    Returns (extract_dir, member_names, notes).
    """
    stem = _archive_stem_for_extract_dir(archive_path)
    dest = supplementary_dir / "_extracted" / stem
    stamp = dest / ".extraction_stamp.json"
    try:
        src_mtime = int(archive_path.stat().st_mtime_ns)
    except OSError:
        src_mtime = 0
    try:
        rel_src = archive_path.relative_to(supplementary_dir).as_posix()
    except ValueError:
        rel_src = archive_path.name
    if stamp.is_file():
        try:
            meta = json.loads(stamp.read_text(encoding="utf-8", errors="ignore"))
            if int(meta.get("mtime_ns", -1)) == src_mtime and meta.get("source") == rel_src:
                manifest = dest / "__members.json"
                if manifest.is_file():
                    data = json.loads(manifest.read_text(encoding="utf-8", errors="ignore"))
                    members = data.get("members") or []
                    if isinstance(members, list):
                        return dest, [str(m) for m in members], ["reused cached extraction"]
        except (json.JSONDecodeError, TypeError, ValueError):
            pass

    dest.mkdir(parents=True, exist_ok=True)
    members, notes = extract_archive_if_needed(archive_path, dest)
    info_paths = find_info_files_in_member_list(members)
    info_snippets: list[str] = []
    for rel in info_paths:
        p = _safe_extract_path(dest, rel)
        if p and p.is_file() and p.stat().st_size <= 512 * 1024:
            try:
                txt = p.read_text(encoding="utf-8", errors="ignore")
                info_snippets.append(f"### info file `{rel}`\n{txt}")
            except OSError:
                pass
    manifest_payload = {
        "source_archive": rel_src,
        "members": members,
        "notes": notes,
        "info_file_previews": info_snippets,
    }
    (dest / "__members.json").write_text(json.dumps(manifest_payload, ensure_ascii=False, indent=2), encoding="utf-8")
    stamp.write_text(
        json.dumps({"source": rel_src, "mtime_ns": src_mtime}, indent=2),
        encoding="utf-8",
    )
    return dest, members, notes


def _safe_extract_path(base: Path, member_name: str) -> Path | None:
    try:
        target = (base / member_name).resolve()
        base_r = base.resolve()
        target.relative_to(base_r)
    except (OSError, ValueError):
        return None
    return target


def _readable_basename_for_supp(path: Path, supplementary_dir: Path) -> str:
    try:
        rel = path.relative_to(supplementary_dir).as_posix()
    except ValueError:
        rel = path.name
    return _safe_name(rel.replace("/", "__"))


def _supplementary_asset_outputs_current(
    path: Path,
    *,
    supplementary_dir: Path,
    readable_dir: Path,
    converted_spreadsheet_dir: Path,
) -> bool:
    """True if an existing conversion output is at least as new as the source (skip re-run)."""
    suffix = path.suffix.lower()
    try:
        src_mtime = path.stat().st_mtime_ns
    except OSError:
        return False
    key = _readable_basename_for_supp(path, supplementary_dir)
    if suffix == ".pdf":
        out_md = readable_dir / f"{key}.md"
        if not out_md.is_file():
            return False
        return out_md.stat().st_mtime_ns >= src_mtime
    if suffix in {".xls", ".xlsx"}:
        base = key
        if not converted_spreadsheet_dir.is_dir():
            return False
        matches = list(converted_spreadsheet_dir.glob(f"{base}__*.csv"))
        if not matches:
            return False
        newest = max(m.stat().st_mtime_ns for m in matches)
        return newest >= src_mtime
    return False


def _convert_excel_to_csvs(
    path: Path, output_dir: Path, *, output_base: str | None = None
) -> tuple[list[Path], list[str]]:
    """Convert xls/xlsx workbook sheets to CSV files (best effort)."""
    csv_paths: list[Path] = []
    notes: list[str] = []
    output_dir.mkdir(parents=True, exist_ok=True)
    base = output_base if output_base is not None else _safe_name(path.stem)

    # First choice: pandas handles xls/xlsx broadly if engines are installed.
    try:
        import pandas as pd  # type: ignore

        sheets = pd.read_excel(path, sheet_name=None)
        for idx, (sheet_name, df) in enumerate(sheets.items(), start=1):
            sheet_slug = _safe_name(str(sheet_name))
            target = output_dir / f"{base}__{idx:02d}_{sheet_slug}.csv"
            df.to_csv(target, index=False)
            csv_paths.append(target)
        if csv_paths:
            notes.append(f"Converted {path.name} to {len(csv_paths)} CSV file(s).")
            return csv_paths, notes
    except Exception as exc:  # noqa: BLE001
        notes.append(f"Pandas conversion failed for {path.name}: {exc}")

    # Fallback for xlsx with openpyxl.
    if path.suffix.lower() == ".xlsx":
        try:
            from openpyxl import load_workbook  # type: ignore

            wb = load_workbook(filename=path, read_only=True, data_only=True)
            for idx, ws in enumerate(wb.worksheets, start=1):
                sheet_slug = _safe_name(ws.title)
                target = output_dir / f"{base}__{idx:02d}_{sheet_slug}.csv"
                with target.open("w", encoding="utf-8", newline="") as f:
                    for row in ws.iter_rows(values_only=True):
                        vals = ["" if v is None else str(v) for v in row]
                        # Minimal CSV escaping.
                        escaped = []
                        for val in vals:
                            v = val.replace('"', '""')
                            if any(ch in v for ch in [",", '"', "\n"]):
                                v = f'"{v}"'
                            escaped.append(v)
                        f.write(",".join(escaped) + "\n")
                csv_paths.append(target)
            if csv_paths:
                notes.append(f"Converted {path.name} to {len(csv_paths)} CSV file(s) via openpyxl.")
                return csv_paths, notes
        except Exception as exc:  # noqa: BLE001
            notes.append(f"Openpyxl conversion failed for {path.name}: {exc}")

    return csv_paths, notes


def _marker_pdf_to_readable_md(path: Path, readable_dir: Path) -> bool:
    """Run convert_pdf.py on a supplementary PDF; write one markdown + extracted images. Returns success."""
    out_md = readable_dir / f"{_readable_basename_for_supp(path, readable_dir.parent)}.md"
    if out_md.exists() and out_md.stat().st_mtime >= path.stat().st_mtime:
        return True
    try:
        readable_dir.mkdir(parents=True, exist_ok=True)
    except OSError:
        return False
    with tempfile.TemporaryDirectory(prefix="supp_pdf_marker_") as tmpdir:
        out_dir = Path(tmpdir)
        # Use convert_pdf.py (which has VRAM tuning) instead of marker_single.
        convert_script = Path(__file__).resolve().parent.parent / "extraction" / "convert_pdf.py"
        supp_name = _readable_basename_for_supp(path, readable_dir.parent)
        command = [
            sys.executable,
            str(convert_script),
            str(path),
            str(out_dir),
            "--name",
            supp_name,
        ]
        completed = subprocess.run(command, capture_output=True, text=True, check=False)
        if completed.returncode != 0:
            return False
        md_path = out_dir / f"{supp_name}.md"
        if not md_path.exists():
            return False
        out_md.write_text(
            md_path.read_text(encoding="utf-8", errors="ignore"),
            encoding="utf-8",
        )
        # Copy extracted images alongside the markdown.
        image_exts = {".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg"}
        img_dir = out_dir / supp_name
        if img_dir.is_dir():
            for img in sorted(img_dir.rglob("*")):
                if img.is_file() and img.suffix.lower() in image_exts:
                    dest = readable_dir / img.name
                    dest.write_bytes(img.read_bytes())
    return True


def prepare_supplementary_for_assessment(*, article_dir: Path, skip_pdf_marker: bool = False) -> str:
    """
    Convert non-text supplementary assets to AI-readable form before assessment.

    - ZIP / tar.gz / tgz: extracted under supplementary/_extracted/<stem>/ (with member manifest)
    - PDF: converted with ``marker_single`` -> ``supplementary/_readable/<key>.md`` (full file on disk),
      unless ``skip_pdf_marker=True`` (spreadsheets only; used by ``tmp_test_example_papers``).
    - .xls/.xlsx: sheet CSVs under supplementary/_converted_spreadsheets/
    """
    supplementary_dir = article_dir / "supplementary"
    if not supplementary_dir.is_dir():
        return "no supplementary directory"

    readable_dir = supplementary_dir / "_readable"
    converted_spreadsheet_dir = supplementary_dir / "_converted_spreadsheets"

    if skip_pdf_marker:
        print("[info] prepare_supplementary: skip_pdf_marker=True (Excel/CSV only; no Marker on PDFs).", flush=True)

    archive_notes: list[str] = []
    for path in sorted(supplementary_dir.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        try:
            rel = path.relative_to(supplementary_dir)
        except ValueError:
            continue
        if "_extracted" in rel.parts or "_readable" in rel.parts or "_converted_spreadsheets" in rel.parts:
            continue
        low = path.name.lower()
        if low.endswith(".zip") or low.endswith(".tar.gz") or low.endswith(".tgz"):
            _dest, members, en = _ensure_archive_extracted(path, supplementary_dir)
            archive_notes.append(f"extract `{path.name}` -> {len(members)} member(s); " + "; ".join(en[:2]))

    pdf_ok = 0
    pdf_fail = 0
    sheet_ok = 0
    sheet_fail = 0

    work_paths: list[Path] = []
    for path in sorted(supplementary_dir.rglob("*")):
        if not path.is_file() or path.name.startswith("."):
            continue
        if path.name == "__members.json":
            continue
        try:
            rel = path.relative_to(supplementary_dir)
        except ValueError:
            continue
        if "_readable" in rel.parts or "_converted_spreadsheets" in rel.parts:
            continue
        suffix = path.suffix.lower()
        if suffix in {".xls", ".xlsx"}:
            work_paths.append(path)
        elif suffix == ".pdf" and not skip_pdf_marker:
            work_paths.append(path)
    n = len(work_paths)
    if n == 0:
        if archive_notes:
            print("[summary] supplementary_conversion: archives only (no PDF/Excel to convert).")
            return "archives_extracted=" + str(len(archive_notes))
        return "no supplementary files to convert"
    if all(
        _supplementary_asset_outputs_current(
            p,
            supplementary_dir=supplementary_dir,
            readable_dir=readable_dir,
            converted_spreadsheet_dir=converted_spreadsheet_dir,
        )
        for p in work_paths
    ):
        print(
            "[info] skip convert_supplementary_to_readable: all sources already have up-to-date "
            "outputs in supplementary/_readable/ and/or _converted_spreadsheets/ (remove them to re-run).",
            flush=True,
        )
        return "skipped: supplementary conversion (all outputs up to date)"

    for i, path in enumerate(work_paths, start=1):
        print(f"format converting {i}/{n} {path.name}", flush=True)
        suffix = path.suffix.lower()
        key = _readable_basename_for_supp(path, supplementary_dir)
        if suffix == ".pdf" and not skip_pdf_marker:
            out_md = readable_dir / f"{key}.md"
            if _marker_pdf_to_readable_md(path, readable_dir):
                pdf_ok += 1
            else:
                pdf_fail += 1
        elif suffix in {".xls", ".xlsx"}:
            csvs, _notes = _convert_excel_to_csvs(path, converted_spreadsheet_dir, output_base=key)
            if csvs:
                sheet_ok += 1
            else:
                sheet_fail += 1

    # Human-oriented line (also returned in the machine-friendly tail below for logs/JSON).
    pdf_part = (
        f"PDF_marker_skipped={skip_pdf_marker}"
        if skip_pdf_marker
        else f"supplementary PDF -> _readable/*.md (marker): {pdf_ok} ok, {pdf_fail} failed"
    )
    print(
        f"[summary] supplementary_conversion: "
        f"archives={len(archive_notes)}; "
        f"{pdf_part}; "
        f"Excel -> _converted_spreadsheets/*.csv: {sheet_ok} ok, {sheet_fail} failed."
    )

    return (
        f"archives={len(archive_notes)} skip_pdf_marker={skip_pdf_marker} "
        f"pdf_marker_ok={pdf_ok} pdf_marker_fail={pdf_fail} "
        f"spreadsheets_csv={sheet_ok} spreadsheet_fail={sheet_fail}"
    )


# ---------------------------------------------------------------------------
# Supplementary file tree for AI-driven data file analysis
# ---------------------------------------------------------------------------


def _human_file_size(n: int) -> str:
    if n < 1024:
        return f"{n} B"
    if n < 1024 * 1024:
        return f"{n / 1024:.1f} KiB"
    return f"{n / (1024 * 1024):.1f} MiB"


def build_supplementary_file_tree(article_dir: Path) -> list[dict]:
    """
    Walk supplementary/ (including _extracted/ archive members) and build a
    list of non-PDF data files with path, size, and type info.
    Skips internal processing dirs (_readable, _converted_spreadsheets).
    """
    supplementary_dir = article_dir / "supplementary"
    if not supplementary_dir.is_dir():
        return []

    tree: list[dict] = []
    seen_rel_paths: set[str] = set()

    for path in sorted(supplementary_dir.rglob("*")):
        if not path.is_file():
            continue
        if path.name.startswith("."):
            continue
        if path.name == AUDIT_FILENAME:
            continue
        try:
            rel = path.relative_to(supplementary_dir).as_posix()
        except ValueError:
            continue
        parts = Path(rel).parts
        # Skip internal conversion dirs
        if "_readable" in parts or "_converted_spreadsheets" in parts:
            continue
        if rel in seen_rel_paths:
            continue
        seen_rel_paths.add(rel)

        suffix = path.suffix.lower()
        is_data_like = suffix in {
            ".csv", ".tsv", ".txt", ".json", ".yaml", ".yml", ".xml",
            ".md", ".dat", ".h5", ".hdf5", ".nc", ".fits", ".root",
            ".npy", ".npz", ".mat",
        }
        try:
            size = path.stat().st_size
        except OSError:
            size = 0

        entry: dict = {
            "path": rel,
            "size_bytes": size,
            "size_human": _human_file_size(size),
            "type": suffix.lstrip(".") if suffix else "unknown",
        }

        if suffix in {".xls", ".xlsx"}:
            # For Excel files, try to get sheet info
            try:
                import pandas as pd
                sheets = pd.read_excel(path, sheet_name=None)
                entry["sheets"] = list(sheets.keys())
                entry["sheet_details"] = []
                for sn, df in sheets.items():
                    entry["sheet_details"].append(
                        f"{sn}: {df.shape[0]} rows x {df.shape[1]} cols"
                    )
            except Exception:
                pass

        tree.append(entry)

    # Also include archive member lists from manifests
    for manifest_path in sorted(supplementary_dir.rglob("__members.json")):
        if "_extracted" not in manifest_path.parts:
            continue
        try:
            manifest = json.loads(manifest_path.read_text(encoding="utf-8", errors="ignore"))
        except (json.JSONDecodeError, OSError):
            continue
        src = manifest.get("source_archive", "?")
        members = manifest.get("members", [])
        member_entries: list[dict] = []
        for m in members:
            p = Path(m)
            ext = p.suffix.lower()
            if ext in {".pdf"}:
                continue
            member_entries.append({
                "member_path": m,
                "type": ext.lstrip(".") if ext else "unknown",
                "is_data_like": ext in {
                    ".csv", ".tsv", ".txt", ".json", ".yaml", ".yml",
                    ".xml", ".dat", ".h5", ".hdf5", ".nc", ".fits",
                },
            })
        if member_entries:
            tree.append({
                "path": f"_extracted/{Path(manifest_path).parent.name}/",
                "type": "archive_members",
                "source_archive": src,
                "members": member_entries,
            })

    return tree


def format_tree_for_ai_prompt(tree: list[dict], *, max_example_items: int = 100) -> str:
    """Format the supplementary file tree as a readable prompt for the AI."""
    if not tree:
        return "(No supplementary data files found.)"

    lines: list[str] = [
        "## Supplementary Data File Tree",
        "",
        "Below is the structure of downloaded supplementary materials (non-PDF files).",
        "",
    ]

    for entry in tree:
        p = entry.get("path", "?")
        sz = entry.get("size_human", "")
        typ = entry.get("type", "")

        if typ == "archive_members":
            src = entry.get("source_archive", "?")
            members = entry.get("members", [])
            lines.append(f"[Archive] {p}  (from {src})")
            shown = 0
            for m in members:
                if shown >= max_example_items:
                    lines.append(f"  ... and {len(members) - shown} more files")
                    break
                mp = m.get("member_path", "?")
                mt = m.get("type", "")
                dl = "(data-like)" if m.get("is_data_like") else ""
                lines.append(f"  {mp}  [{mt}] {dl}")
                shown += 1
            lines.append("")
            continue

        dl_note = ""
        if typ in {"csv", "tsv", "txt", "json", "dat"}:
            dl_note = " [data-like]"

        sheet_info = ""
        sheets = entry.get("sheet_details")
        if sheets:
            sheet_info = "  Sheets: " + "; ".join(sheets)

        lines.append(f"  {sz:>10}  {p}  [{typ}]{dl_note}")
        if sheet_info:
            lines.append(f"           {sheet_info}")

    return "\n".join(lines)


# ---------------------------------------------------------------------------
# PaperContext
# ---------------------------------------------------------------------------


@dataclass(slots=True)
class PaperContext:
    metadata: dict
    primary_markdown: str
    converted_markdown: str
    supplementary_listing: list[str]
    supplementary_text_snippets: list[str]
    source_notes: list[str]
    image_paths: list[Path]
    source_data_audit_markdown: str = ""
    data_file_analysis_overview: str = ""
    data_file_analysis_selected: list[dict] = field(default_factory=list)
    data_file_analysis_snippets: list[str] = field(default_factory=list)
    figure_table_assessment_block: str = ""
    figure_table_final_data_block: str = ""
    derivation_assessment_block: str = ""
    derivation_assessment_final_block: str = ""
    pipeline_assessment_block: str = ""
    pipeline_assessment_final_block: str = ""

    def to_prompt_block(self) -> str:
        metadata_block = "\n".join(f"- {k}: {v}" for k, v in sorted(self.metadata.items()))
        supp_files = "\n".join(f"- {name}" for name in self.supplementary_listing) or "- none"
        supp_snippets = "\n\n".join(self.supplementary_text_snippets) or "No supplementary text snippets."
        notes = "\n".join(f"- {line}" for line in self.source_notes) or "- none"
        images = "\n".join(f"- {p.name}" for p in self.image_paths) or "- none"
        audit = (self.source_data_audit_markdown or "").strip()
        audit_block = f"\n{audit}\n" if audit else ""
        data_overview = (self.data_file_analysis_overview or "").strip()
        data_snippets = "\n\n".join(self.data_file_analysis_snippets) if self.data_file_analysis_snippets else ""
        # The overview block already contains its own section headers
        data_overview_block = f"\n\n{data_overview}\n" if data_overview else ""
        data_snippets_block = (
            f"\n{data_snippets}\n"
            if data_snippets else ""
        )
        ft_block = (self.figure_table_assessment_block or "").strip()
        ft_block_formatted = f"\n\n{ft_block}\n" if ft_block else ""
        fd_block = (self.figure_table_final_data_block or "").strip()
        fd_block_formatted = f"\n\n{fd_block}\n" if fd_block else ""
        da_block = (self.derivation_assessment_block or "").strip()
        da_block_formatted = f"\n\n{da_block}\n" if da_block else ""
        da_final_block = (self.derivation_assessment_final_block or "").strip()
        da_final_block_formatted = f"\n\n{da_final_block}\n" if da_final_block else ""
        pa_block = (self.pipeline_assessment_block or "").strip()
        pa_block_formatted = f"\n\n{pa_block}\n" if pa_block else ""
        pa_final_block = (self.pipeline_assessment_final_block or "").strip()
        pa_final_block_formatted = f"\n\n{pa_final_block}\n" if pa_final_block else ""
        return (
            "## Paper Metadata\n"
            f"{metadata_block}\n\n"
            "## Primary Article Markdown\n"
            f"{self.primary_markdown}\n\n"
            "## Converted PDF Markdown\n"
            f"{self.converted_markdown}\n\n"
            "## Supplementary Files\n"
            f"{supp_files}\n\n"
            "## Supplementary Snippets\n"
            f"{supp_snippets}\n\n"
            "## Images Found\n"
            f"{images}\n\n"
            "## Source Notes\n"
            f"{notes}\n"
            f"{audit_block}"
            f"{data_overview_block}"
            f"{data_snippets_block}"
            f"{ft_block_formatted}"
            f"{fd_block_formatted}"
            f"{da_block_formatted}"
            f"{da_final_block_formatted}"
            f"{pa_block_formatted}"
            f"{pa_final_block_formatted}"
        )


def build_paper_context(
    *,
    article_dir: Path,
    metadata: dict,
    max_primary_chars: int | None = None,
    max_converted_chars: int | None = MAX_CONVERTED_CHARS,
    max_supp_snippet_chars: int = 32000,
    max_images: int = 10,
) -> PaperContext:
    article_md_path = article_dir / "article.md"
    converted_md_path = article_dir / "converted" / "article.md"
    supplementary_dir = article_dir / "supplementary"
    converted_dir = article_dir / "converted"
    source_notes: list[str] = []

    primary_markdown = ""
    if article_md_path.exists():
        primary_markdown = article_md_path.read_text(encoding="utf-8", errors="ignore")
        source_notes.append(f"Primary markdown loaded from {article_md_path.name}.")
    else:
        source_notes.append("Primary markdown missing.")
    primary_markdown = _truncate(primary_markdown, max_primary_chars)

    converted_markdown = ""
    if converted_md_path.exists():
        converted_markdown = converted_md_path.read_text(encoding="utf-8", errors="ignore")
        source_notes.append(f"Converted markdown loaded from {converted_md_path.as_posix()}.")
    else:
        source_notes.append("Converted PDF markdown missing.")
    converted_markdown = _truncate(converted_markdown, max_converted_chars)

    supplementary_listing: list[str] = []
    supplementary_text_snippets: list[str] = []
    if supplementary_dir.exists():
        text_like_suffixes = {".md", ".txt", ".csv", ".tsv", ".json", ".yaml", ".yml", ".xml"}
        converted_spreadsheet_dir = supplementary_dir / "_converted_spreadsheets"
        for path in sorted(supplementary_dir.rglob("*")):
            if not path.is_file():
                continue
            if path.name.startswith("."):
                continue
            if path.name == AUDIT_FILENAME:
                continue
            if "_converted_spreadsheets" in path.parts:
                continue
            if "_readable" in path.parts:
                # Marker output for supplementary PDFs; those files are inlined from the .pdf entry (full text, no duplicate).
                continue
            try:
                rel_name = path.relative_to(supplementary_dir).as_posix()
            except Exception:
                rel_name = path.name
            supplementary_listing.append(rel_name)
            suffix = path.suffix.lower()
            if suffix in text_like_suffixes:
                body = _text_file_snippet_for_model(
                    path,
                    max_full=SUPP_TEXT_FULL_MAX_CHARS,
                    max_preview=max(SUPP_TEXT_PREVIEW_MAX_CHARS, min(max_supp_snippet_chars, 48_000)),
                    read_cap_bytes=SUPP_TEXT_READ_CAP_BYTES,
                )
                if body.strip():
                    supplementary_text_snippets.append(
                        f"### {rel_name}\n{_truncate(body, max_supp_snippet_chars)}"
                    )
            elif suffix == ".pdf":
                key = _readable_basename_for_supp(path, supplementary_dir)
                pre_md = supplementary_dir / "_readable" / f"{key}.md"
                if pre_md.is_file():
                    pre_text = pre_md.read_text(encoding="utf-8", errors="ignore")
                    if pre_text.strip():
                        supplementary_text_snippets.append(
                            f"### {rel_name} (from marker; full Marker output)\n{pre_text}"
                        )
                        source_notes.append(f"Supplementary PDF used pre-converted: {key}.md (full text in prompt)")
                    else:
                        pdf_text = _extract_pdf_text_snippet(path, None)
                        if pdf_text.strip():
                            supplementary_text_snippets.append(
                                f"### {rel_name} (Marker or pypdf fallback; full extract)\n{pdf_text}"
                            )
                else:
                    pdf_text = _extract_pdf_text_snippet(path, None)
                    if pdf_text.strip():
                        supplementary_text_snippets.append(
                            f"### {rel_name} (Marker or pypdf fallback; full extract)\n{pdf_text}"
                        )
            elif suffix in {".xls", ".xlsx"}:
                base = _readable_basename_for_supp(path, supplementary_dir)
                existing_csvs = (
                    sorted(converted_spreadsheet_dir.glob(f"{base}__*.csv")) if converted_spreadsheet_dir.is_dir() else []
                )
                if existing_csvs:
                    csv_files = existing_csvs
                    conversion_notes = [f"Reused {len(existing_csvs)} pre-converted CSV for {path.name}."]
                else:
                    csv_files, conversion_notes = _convert_excel_to_csvs(
                        path, converted_spreadsheet_dir, output_base=base
                    )
                for n in conversion_notes:
                    source_notes.append(f"Spreadsheet conversion: {n}")
                for csv_path in csv_files:
                    try:
                        csv_rel = csv_path.relative_to(supplementary_dir).as_posix()
                    except Exception:
                        csv_rel = csv_path.name
                    supplementary_listing.append(csv_rel)
                    csv_body = _text_file_snippet_for_model(
                        csv_path,
                        max_full=SUPP_TEXT_FULL_MAX_CHARS,
                        max_preview=max(SUPP_TEXT_PREVIEW_MAX_CHARS, min(max_supp_snippet_chars, 48_000)),
                        read_cap_bytes=SUPP_TEXT_READ_CAP_BYTES,
                    )
                    if csv_body.strip():
                        supplementary_text_snippets.append(
                            f"### {csv_rel}\n{_truncate(csv_body, max_supp_snippet_chars)}"
                        )
                excel_text, excel_notes = _extract_excel_text_snippet(path, max_supp_snippet_chars)
                if excel_text.strip():
                    supplementary_text_snippets.append(f"### {rel_name}\n{excel_text}")
                for note in excel_notes:
                    source_notes.append(f"Spreadsheet stats: {note}")
        for manifest in sorted(supplementary_dir.rglob("__members.json")):
            if "_extracted" not in manifest.parts:
                continue
            try:
                mrel = manifest.relative_to(supplementary_dir).as_posix()
            except ValueError:
                mrel = manifest.name
            try:
                mtxt = manifest.read_text(encoding="utf-8", errors="ignore")
            except OSError:
                continue
            if mtxt.strip():
                supplementary_text_snippets.append(
                    f"### Archive manifest {mrel}\n{_truncate(mtxt, max_supp_snippet_chars)}"
                )
    if not supplementary_listing:
        source_notes.append("No supplementary files were found.")

    # Enforce total supplementary snippet budget to stay within model context limit.
    if supplementary_text_snippets:
        supp_chars = sum(len(s) for s in supplementary_text_snippets)
        if supp_chars > MAX_SUPP_SNIPPET_TOTAL_CHARS:
            budget = MAX_SUPP_SNIPPET_TOTAL_CHARS
            kept: list[str] = []
            for s in supplementary_text_snippets:
                if len(s) <= budget:
                    kept.append(s)
                    budget -= len(s)
                else:
                    kept.append(s[:budget])
                    kept.append(
                        f"\n\n[... supplementary snippet budget reached at "
                        f"{MAX_SUPP_SNIPPET_TOTAL_CHARS} chars; further snippets omitted]"
                    )
                    source_notes.append(
                        f"Supplementary snippet total {supp_chars} chars exceeded "
                        f"budget {MAX_SUPP_SNIPPET_TOTAL_CHARS}; truncated."
                    )
                    break
            supplementary_text_snippets = kept

    statement_scan = scan_availability_statement(load_combined_article_markdown(article_dir))
    audit_payload = read_audit_file(supplementary_dir) if supplementary_dir.exists() else None
    source_data_audit_markdown = format_audit_for_prompt(audit_payload, statement_rescan=statement_scan)

    image_paths: list[Path] = []
    if converted_dir.exists() or supplementary_dir.exists():
        image_extensions = {".png", ".jpg", ".jpeg", ".webp"}
        all_images = []

        # Collect images from main article Marker output.
        if converted_dir.exists():
            for p in converted_dir.rglob("*"):
                if p.is_file() and p.suffix.lower() in image_extensions:
                    all_images.append(p)

        # Collect images from supplementary PDF Marker output.
        supp_readable = supplementary_dir / "_readable" if supplementary_dir.exists() else None
        if supp_readable and supp_readable.is_dir():
            for p in supp_readable.rglob("*"):
                if p.is_file() and p.suffix.lower() in image_extensions:
                    all_images.append(p)

        # Sort and limit images (main article images first, then supplementary).
        all_images.sort(key=lambda x: (0 if "converted" in str(x) else 1, x.name))
        image_paths = all_images[:max_images]
        if image_paths:
            main_n = sum(1 for p in image_paths if "converted" in str(p))
            supp_n = len(image_paths) - main_n
            parts = []
            if main_n:
                parts.append(f"{main_n} from main article")
            if supp_n:
                parts.append(f"{supp_n} from supplementary PDFs")
            source_notes.append(f"Loaded {len(image_paths)} images for visual analysis ({', '.join(parts)}).")

    return PaperContext(
        metadata=metadata,
        primary_markdown=primary_markdown,
        converted_markdown=converted_markdown,
        supplementary_listing=supplementary_listing,
        supplementary_text_snippets=supplementary_text_snippets,
        source_notes=source_notes,
        image_paths=image_paths,
        source_data_audit_markdown=source_data_audit_markdown,
    )

